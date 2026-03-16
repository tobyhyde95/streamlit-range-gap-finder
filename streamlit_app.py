#!/usr/bin/env python3
"""
Range Gap Finder - Streamlit application.

Runs Content Gaps, Competitive Opportunities, Market Share,
and Product Opportunity Group analysis.

Pipeline: De-duplication → Stock Classification → Semantic Clustering → Scoring → Report
"""

import io
import json
import os
import sys
import tempfile
import uuid

import pandas as pd
import streamlit as st

# Add project root so seo_analyzer can be imported
PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

from seo_analyzer.data_loader import read_csv_with_encoding_fallback


def _infer_columns(df: pd.DataFrame) -> dict:
    """Suggest column mapping from dataframe columns."""
    cols = list(df.columns)
    lower = [c.lower() for c in cols]
    mapping = {}
    if cols:
        for name in ("keyword", "keywords", "query", "queries"):
            for i, c in enumerate(lower):
                if name in c and "col" not in c:
                    mapping["keywordCol"] = cols[i]
                    break
            if "keywordCol" in mapping:
                break
        if "keywordCol" not in mapping and cols:
            mapping["keywordCol"] = cols[0]
        for name in ("position", "rank", "ranking", "pos"):
            for i, c in enumerate(lower):
                if name in c:
                    mapping["positionCol"] = cols[i]
                    break
            if "positionCol" in mapping:
                break
        if "positionCol" not in mapping and len(cols) > 1:
            mapping["positionCol"] = cols[1]
        for i, c in enumerate(lower):
            if "url" in c or "landing" in c:
                mapping["urlCol"] = cols[i]
                break
        if "urlCol" not in mapping and len(cols) > 2:
            mapping["urlCol"] = cols[2]
        for name in ("volume", "search volume", "searches", "search volume"):
            for i, c in enumerate(lower):
                if "volume" in c or "search" in c and "volume" in c or c == "searches":
                    mapping["volumeCol"] = cols[i]
                    break
            if "volumeCol" in mapping:
                break
        for name in ("traffic", "organic traffic", "traffic value"):
            for i, c in enumerate(lower):
                if "traffic" in c:
                    mapping["trafficCol"] = cols[i]
                    break
    return mapping


def _get_csv_columns(uploaded_file) -> list:
    """Read first CSV and return column names."""
    uploaded_file.seek(0)
    df = read_csv_with_encoding_fallback(uploaded_file)
    if df is not None and not df.columns.empty:
        return list(df.columns)
    uploaded_file.seek(0)
    try:
        df = pd.read_csv(uploaded_file, nrows=1)
        return list(df.columns)
    except Exception:
        return []


def _color_score(val):
    """Conditional color-coding for Opportunity Scores (NFR1)."""
    try:
        score = float(val)
    except (ValueError, TypeError):
        return ""
    if score >= 80:
        return "background-color: #27ae60; color: white; font-weight: bold"
    elif score >= 50:
        return "background-color: #f39c12; color: white; font-weight: bold"
    else:
        return "background-color: #e74c3c; color: white; font-weight: bold"


def _to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Report") -> bytes:
    """Convert a DataFrame to Excel bytes with conditional formatting."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)

        # Apply conditional formatting to Opportunity Score column
        if "Opportunity Score" in df.columns:
            try:
                from openpyxl.formatting.rule import CellIsRule
                from openpyxl.styles import PatternFill

                ws = writer.sheets[sheet_name]
                score_col_idx = list(df.columns).index("Opportunity Score") + 1
                col_letter = chr(64 + score_col_idx) if score_col_idx <= 26 else "A"

                green_fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
                amber_fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
                red_fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")

                cell_range = f"{col_letter}2:{col_letter}{len(df) + 1}"

                ws.conditional_formatting.add(cell_range,
                    CellIsRule(operator="greaterThanOrEqual", formula=["80"], fill=green_fill))
                ws.conditional_formatting.add(cell_range,
                    CellIsRule(operator="between", formula=["50", "79"], fill=amber_fill))
                ws.conditional_formatting.add(cell_range,
                    CellIsRule(operator="lessThan", formula=["50"], fill=red_fill))
            except ImportError:
                pass

    return buf.getvalue()


def run_analysis(our_path, competitor_paths, onsite_path, options_str, status_widget, catalogue_path=None):
    """Run full analysis. Imports services here so the app starts fast on Streamlit Cloud."""
    from seo_analyzer import services

    def report(message, current, total):
        if status_widget is not None:
            try:
                status_widget.update(label=message, state="running")
            except Exception:
                pass

    return services.run_full_analysis(
        our_file_path=our_path,
        competitor_file_paths=competitor_paths or [],
        onsite_file_path=onsite_path,
        options_str=options_str,
        progress_reporter=report,
        catalogue_file_path=catalogue_path,
    )


def main():
    st.set_page_config(
        page_title="Range Gap Finder",
        page_icon="🔍",
        layout="wide",
        initial_sidebar_state="expanded",
    )
    st.title("🔍 Intelligent Gap Analysis Tool")
    st.caption("Product Opportunity Groups · Content Gaps · Competitive Opportunities · Market Share")

    # Sidebar: file uploads
    with st.sidebar:
        st.header("Data Sources")
        our_file = st.file_uploader(
            "Ahrefs domain export (CSV)",
            type=["csv"],
            key="our_file",
            help="Your site's Ahrefs keyword export (keyword, position, URL, volume, traffic).",
        )
        competitor_files = st.file_uploader(
            "Competitor Ahrefs exports (CSV)",
            type=["csv"],
            accept_multiple_files=True,
            key="competitor_files",
            help="One or more competitor Ahrefs CSV exports.",
        )
        onsite_file = st.file_uploader(
            "GA4 on-site search data (optional)",
            type=["csv"],
            key="onsite_file",
            help="GA4 internal site search CSV with keyword + search count columns.",
        )
        catalogue_file = st.file_uploader(
            "Product catalogue (optional — enables stock classification)",
            type=["csv"],
            key="catalogue_file",
            help="Product catalogue CSV. Auto-detects product name column. Enables Stocked/Not Stocked classification.",
        )

        st.divider()
        st.header("Filters (optional)")
        rank_from = st.number_input("Rank from", min_value=1, value=None, placeholder="e.g. 1")
        rank_to = st.number_input("Rank to", min_value=1, value=None, placeholder="e.g. 100")
        excluded_keywords_text = st.text_area(
            "Excluded keywords (one per line)",
            height=100,
            placeholder="brand1\nbrand2",
            help="Keywords containing these terms will be excluded.",
        )
        excluded_keywords = [x.strip() for x in (excluded_keywords_text or "").splitlines() if x.strip()]

    # Column mapping (after at least our file is uploaded)
    column_map = {}
    if our_file:
        cols = _get_csv_columns(our_file)
        if cols:
            inferred = _infer_columns(pd.DataFrame(columns=cols))
            st.subheader("Column mapping")
            c1, c2, c3 = st.columns(3)

            def idx(opt, default=0):
                return cols.index(opt) if opt in cols else default

            with c1:
                column_map["keywordCol"] = st.selectbox("Keyword column", options=cols, index=idx(inferred.get("keywordCol"), 0), key="kw")
                column_map["positionCol"] = st.selectbox("Position / Rank column", options=cols, index=idx(inferred.get("positionCol"), min(1, len(cols) - 1)), key="pos")
            with c2:
                column_map["urlCol"] = st.selectbox("URL column", options=cols, index=idx(inferred.get("urlCol"), min(2, len(cols) - 1)), key="url")
                column_map["volumeCol"] = st.selectbox("Volume column (Ahrefs monthly searches)", options=[""] + cols, key="vol")
                if column_map["volumeCol"] == "":
                    column_map["volumeCol"] = None
            with c3:
                column_map["trafficCol"] = st.selectbox("Traffic column (optional)", options=[""] + cols, key="traffic")
                if column_map["trafficCol"] == "":
                    column_map["trafficCol"] = None
            if column_map.get("volumeCol") is None:
                column_map.pop("volumeCol", None)
            if column_map.get("trafficCol") is None:
                column_map.pop("trafficCol", None)

    # Lenses to run
    st.subheader("Lenses to run")
    lens_c1, lens_c2, lens_c3 = st.columns(3)
    with lens_c1:
        run_content_gaps = st.checkbox("Content Gaps", value=True, help="Keywords competitors rank for but you don't")
    with lens_c2:
        run_competitive_opportunities = st.checkbox("Competitive Opportunities", value=True, help="Keywords where competitors outrank you")
    with lens_c3:
        run_market_share = st.checkbox("Market Share", value=True, help="Traffic distribution across domains")

    run_analysis_btn = st.button("Run analysis", type="primary", use_container_width=True)

    if not run_analysis_btn:
        st.info("Upload your CSV(s) and set column mapping, then click **Run analysis**.")
        return

    if not our_file:
        st.error("Please upload your Ahrefs domain export (CSV).")
        return

    if not column_map or not column_map.get("keywordCol") or not column_map.get("positionCol") or not column_map.get("urlCol"):
        st.error("Please set Keyword, Position, and URL columns.")
        return

    # Build options JSON
    options = {
        "columnMap": {
            "keywordCol": column_map["keywordCol"],
            "positionCol": column_map["positionCol"],
            "urlCol": column_map["urlCol"],
            "volumeCol": column_map.get("volumeCol") or "",
            "trafficCol": column_map.get("trafficCol") or "",
        },
        "lensesToRun": {
            "content_gaps": run_content_gaps,
            "competitive_opportunities": run_competitive_opportunities,
            "market_share": run_market_share,
        },
        "rankFrom": int(rank_from) if rank_from is not None else None,
        "rankTo": int(rank_to) if rank_to is not None else None,
        "excludedKeywords": excluded_keywords,
        "onsiteDateRange": "",
    }

    # Save uploads to temp dir and run
    progress_placeholder = st.empty()
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            our_path = os.path.join(temp_dir, f"{uuid.uuid4()}.csv")
            with open(our_path, "wb") as f:
                f.write(our_file.getvalue())

            comp_paths = []
            for cf in (competitor_files or []):
                p = os.path.join(temp_dir, f"{uuid.uuid4()}.csv")
                with open(p, "wb") as f:
                    f.write(cf.getvalue())
                comp_paths.append(p)

            onsite_path = None
            if onsite_file:
                onsite_path = os.path.join(temp_dir, f"{uuid.uuid4()}.csv")
                with open(onsite_path, "wb") as f:
                    f.write(onsite_file.getvalue())

            catalogue_path = None
            if catalogue_file:
                catalogue_path = os.path.join(temp_dir, f"{uuid.uuid4()}.csv")
                with open(catalogue_path, "wb") as f:
                    f.write(catalogue_file.getvalue())

            with progress_placeholder.container():
                status = st.status("Running analysis…", state="running")
                result = run_analysis(
                    our_path,
                    comp_paths,
                    onsite_path,
                    json.dumps(options),
                    status,
                    catalogue_path=catalogue_path,
                )
                status.update(label="Analysis complete.", state="complete")
    except Exception as e:
        progress_placeholder.empty()
        st.exception(e)
        return

    progress_placeholder.empty()

    # Results
    st.success("Analysis finished. Explore the tabs below.")

    tab_names = ["Product Opportunity Groups", "Content Gaps", "Competitive Opportunities", "Market Share"]
    tabs = st.tabs(tab_names)

    # --- Tab 0: Product Opportunity Groups (new primary output) ---
    with tabs[0]:
        if result.get("productOpportunityReport"):
            st.subheader("Product Opportunity Groups — Ranked by Opportunity Score")
            st.caption(
                "Each row = one Product Opportunity Group (not an individual keyword). "
                "Score = (Normalised Ahrefs Vol × 0.6) + (Normalised Internal Searches × 0.4). "
                "Green ≥80 | Amber 50–79 | Red <50"
            )

            df_opp = pd.DataFrame(result["productOpportunityReport"])

            # Apply conditional color-coding (NFR1)
            if "Opportunity Score" in df_opp.columns:
                styled_df = df_opp.style.applymap(
                    _color_score, subset=["Opportunity Score"]
                )
                st.dataframe(styled_df, use_container_width=True, hide_index=True)
            else:
                st.dataframe(df_opp, use_container_width=True, hide_index=True)

            # Excel download with conditional formatting
            col_dl1, col_dl2 = st.columns(2)
            with col_dl1:
                excel_bytes = _to_excel_bytes(df_opp, sheet_name="Product Opportunity Groups")
                st.download_button(
                    "Download Excel (.xlsx)",
                    data=excel_bytes,
                    file_name="product_opportunity_groups.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_opp_xlsx",
                )
            with col_dl2:
                csv_buf = io.BytesIO()
                df_opp.to_csv(csv_buf, index=False)
                st.download_button(
                    "Download CSV",
                    data=csv_buf.getvalue(),
                    file_name="product_opportunity_groups.csv",
                    mime="text/csv",
                    key="dl_opp_csv",
                )

            # Summary metrics
            if len(df_opp) > 0:
                m1, m2, m3, m4 = st.columns(4)
                m1.metric("Total Groups", len(df_opp))
                high_score = len(df_opp[df_opp.get("Opportunity Score", pd.Series()) >= 80]) if "Opportunity Score" in df_opp.columns else 0
                m2.metric("High Priority (≥80)", high_score)
                total_variants = df_opp["Keyword Variant Count"].sum() if "Keyword Variant Count" in df_opp.columns else 0
                m3.metric("Total Keyword Variants", int(total_variants))
                if result.get("hasCatalogue"):
                    m4.metric("Stock Classification", "Active")
                else:
                    m4.metric("Stock Classification", "No catalogue uploaded")
        else:
            st.info("No Product Opportunity Groups generated. Ensure at least one lens is enabled and data is uploaded.")

    # --- Tab 1: Content Gaps ---
    with tabs[1]:
        if result.get("keywordGapReport"):
            st.subheader("Content Gaps · Individual keywords")
            df_gap = pd.DataFrame(result["keywordGapReport"])
            st.dataframe(df_gap, use_container_width=True, hide_index=True)
            buf = io.BytesIO()
            df_gap.to_csv(buf, index=False)
            st.download_button("Download CSV", data=buf.getvalue(), file_name="content_gaps_keywords.csv", mime="text/csv", key="dl_gap_kw")
        else:
            st.info("No content gaps data. Enable **Content Gaps** and ensure competitor files are uploaded.")
        if result.get("topicGapReport"):
            st.subheader("Content Gaps · Keyword groups (full)")
            df_topic = pd.DataFrame(result["topicGapReport"])
            if "Opportunity Score" in df_topic.columns:
                styled = df_topic.style.applymap(_color_score, subset=["Opportunity Score"])
                st.dataframe(styled, use_container_width=True, hide_index=True)
            else:
                st.dataframe(df_topic, use_container_width=True, hide_index=True)
            buf = io.BytesIO()
            df_topic.to_csv(buf, index=False)
            st.download_button("Download CSV", data=buf.getvalue(), file_name="content_gaps_groups_full.csv", mime="text/csv", key="dl_gap_grp_full")
        if result.get("coreTopicGapReport"):
            st.subheader("Content Gaps · Keyword groups (core, rank ≤20)")
            df_core = pd.DataFrame(result["coreTopicGapReport"])
            if "Opportunity Score" in df_core.columns:
                styled = df_core.style.applymap(_color_score, subset=["Opportunity Score"])
                st.dataframe(styled, use_container_width=True, hide_index=True)
            else:
                st.dataframe(df_core, use_container_width=True, hide_index=True)
            buf = io.BytesIO()
            df_core.to_csv(buf, index=False)
            st.download_button("Download CSV", data=buf.getvalue(), file_name="content_gaps_groups_core.csv", mime="text/csv", key="dl_gap_grp_core")

    # --- Tab 2: Competitive Opportunities ---
    with tabs[2]:
        if result.get("keywordThreatsReport"):
            st.subheader("Competitive Opportunities · Individual keywords")
            df_threat = pd.DataFrame(result["keywordThreatsReport"])
            st.dataframe(df_threat, use_container_width=True, hide_index=True)
            buf = io.BytesIO()
            df_threat.to_csv(buf, index=False)
            st.download_button("Download CSV", data=buf.getvalue(), file_name="competitive_opportunities_keywords.csv", mime="text/csv", key="dl_threat_kw")
        else:
            st.info("No competitive opportunities data. Enable **Competitive Opportunities** and upload competitor files.")
        if result.get("topicThreatsReport"):
            st.subheader("Competitive Opportunities · Keyword groups (full)")
            df_tt = pd.DataFrame(result["topicThreatsReport"])
            st.dataframe(df_tt, use_container_width=True, hide_index=True)
            buf = io.BytesIO()
            df_tt.to_csv(buf, index=False)
            st.download_button("Download CSV", data=buf.getvalue(), file_name="competitive_opportunities_groups_full.csv", mime="text/csv", key="dl_threat_grp_full")
        if result.get("coreTopicThreatsReport"):
            st.subheader("Competitive Opportunities · Keyword groups (core)")
            df_ct = pd.DataFrame(result["coreTopicThreatsReport"])
            st.dataframe(df_ct, use_container_width=True, hide_index=True)
            buf = io.BytesIO()
            df_ct.to_csv(buf, index=False)
            st.download_button("Download CSV", data=buf.getvalue(), file_name="competitive_opportunities_groups_core.csv", mime="text/csv", key="dl_threat_grp_core")

    # --- Tab 3: Market Share ---
    with tabs[3]:
        if result.get("keywordMarketShareReport"):
            st.subheader("Market Share · Individual keywords")
            df_ms = pd.DataFrame(result["keywordMarketShareReport"])
            st.dataframe(df_ms, use_container_width=True, hide_index=True)
            buf = io.BytesIO()
            df_ms.to_csv(buf, index=False)
            st.download_button("Download CSV", data=buf.getvalue(), file_name="market_share_keywords.csv", mime="text/csv", key="dl_ms_kw")
        else:
            st.info("No market share data. Enable **Market Share** and ensure a traffic column is mapped.")
        if result.get("groupMarketShareReport"):
            st.subheader("Market Share · Keyword groups (full)")
            df_gms = pd.DataFrame(result["groupMarketShareReport"])
            st.dataframe(df_gms, use_container_width=True, hide_index=True)
            buf = io.BytesIO()
            df_gms.to_csv(buf, index=False)
            st.download_button("Download CSV", data=buf.getvalue(), file_name="market_share_groups_full.csv", mime="text/csv", key="dl_ms_grp_full")
        if result.get("coreGroupMarketShareReport"):
            st.subheader("Market Share · Keyword groups (core)")
            df_cgms = pd.DataFrame(result["coreGroupMarketShareReport"])
            st.dataframe(df_cgms, use_container_width=True, hide_index=True)
            buf = io.BytesIO()
            df_cgms.to_csv(buf, index=False)
            st.download_button("Download CSV", data=buf.getvalue(), file_name="market_share_groups_core.csv", mime="text/csv", key="dl_ms_grp_core")


if __name__ == "__main__":
    main()
